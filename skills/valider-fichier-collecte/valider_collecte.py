#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
valider_collecte.py  --  Couche C1, deterministe (Python pur, sans LLM).

Controle qu'une trame de collecte FDES remplie par le client est exploitable
avant import SimaPro. Ne corrige rien : signale. Reproductible et auditable
pour la revue critique ISO. Chaque execution logge dans _logs/.

Usage :
    python valider_collecte.py <chemin_trame.xlsx> [<dossier_sortie>] [--tol 0.02]

Sorties :
    <dossier_sortie>/alertes-validation.md
    <dossier_sortie>/_logs/valider-fichier-collecte_<timestamp>.json
Code retour processus : 0 = OK, 2 = BLOQUANT, 1 = erreur technique.
"""

import sys
import os
import json
import argparse
from datetime import datetime, timezone

import openpyxl

# ---------------------------------------------------------------------------
# CONFIGURATION FIGEE  --  structure attendue, derivee de la sortie de
# /generer-trame-collecte. A maintenir ici si la trame evolue (vague suivante).
# ---------------------------------------------------------------------------

# Nom exact de l'onglet bilan massique (cache). Si le contenu est titre
# "9. Bilan Massique" mais l'onglet s'appelle "Bilan Massique", c'est ce
# dernier qui compte ici.
ONGLET_BILAN = "Bilan Massique"

# Tous les onglets attendus dans la trame.
ONGLETS_ATTENDUS = [
    "0. Mode d'emploi",
    "1. Societe et site",          # accent gere plus bas par normalisation souple
    "2. Vos produits",
    "3. Composition des produits",
    "4. Energie - Eau",
    "5. Dechets",
    "6. Rejets air-eau (GEREP)",
    "7. Emballages",
    "8. Livraison et fin de vie",
    ONGLET_BILAN,
]

# En-tetes attendus pour les onglets tabulaires (ligne d'en-tete, colonne de
# depart, libelles). Sert au controle de structure (libelle present == OK).
ENTETES_ATTENDUS = {
    "2. Vos produits": {
        "ligne": 4, "col_depart": 1,
        "libelles": [
            "Reference interne / code", "Nom commercial", "Description / usage",
            "Longueur (mm)", "Largeur (mm)", "Epaisseur / hauteur (mm)",
            "Masse d'une unite (kg)", "Unite de vente",
            "Quantite fabriquee sur l'annee", "Unite de cette quantite",
            "Commentaire",
        ],
    },
    "3. Composition des produits": {
        "ligne": 5, "col_depart": 1,
        "libelles": [
            "Reference produit", "Matiere / composant",
            "Masse pour 1 unite de produit (kg)", "Matiere recyclee ?",
            "Fournisseur", "Ville du fournisseur", "Pays",
            "Source de l'information", "Commentaire",
        ],
    },
    "4. Energie - Eau": {
        "ligne": 4, "col_depart": 2,
        "libelles": ["Type d'energie", "Valeur", "Unite", "Source", "Commentaire"],
    },
    "5. Dechets": {
        "ligne": 4, "col_depart": 1,
        "libelles": [
            "Type de dechet", "Dangereux ?", "Quantite", "Unite",
            "Mode d'elimination / valorisation", "Source", "Commentaire",
        ],
    },
    "6. Rejets air-eau (GEREP)": {
        "ligne": 8, "col_depart": 2,
        "libelles": ["Polluant", "Valeur", "Unite", "Source / annee GEREP", "Commentaire"],
    },
    "7. Emballages": {
        "ligne": 4, "col_depart": 2,
        "libelles": ["Element", "Valeur", "Unite / precision", "Source", "Commentaire"],
    },
}

# Reperes du bilan massique : libelle de garde -> cellule porteuse du total.
BILAN_REPERES = {
    "TOTAL ENTRANTS": {"cellule_libelle": "B12", "cellule_valeur": "C12"},
    "TOTAL SORTANTS": {"cellule_libelle": "B20", "cellule_valeur": "C20"},
}

# ---------------------------------------------------------------------------
# A COMPLETER PAR LE PRATICIEN  --  decisions de methodologie, PAS devinees.
# Tant que ces dicts sont vides, les controles correspondants ne s'executent
# pas (le skill ne fabrique aucune regle). Documenter chaque ajout.
# ---------------------------------------------------------------------------

# Champs obligatoires : onglet -> liste de (libelle_ligne, cellule_valeur).
# Ex. ("Nom de la societe", "C4"). DUVP volontairement absente (decision actee).
CHAMPS_OBLIGATOIRES = {
     "1. Société et site": [
        ("Nom de la société", "C4"),
        ("Personne référente (nom)", "C5"),
        ("Email du référent", "C6"),
        ("Téléphone du référent", "C7"),
        ("Nom du site", "C10"),
        ("Adresse du site", "C11"),
        ("Année de référence des données", "C16"),
        ("Date de remplissage", "C19"),
    ],

    # Tableau répétable : 1 ligne par produit, à partir de la ligne 5
    "2. Vos produits": [
        ("Référence interne / code", "A5"),
        ("Nom commercial", "B5"),
        ("Masse d'une unité (kg)", "G5"),
        ("Unité de vente", "H5"),
        ("Quantité fabriquée sur l'année", "I5"),
        ("Unité de cette quantité", "J5"),
    ],

    # Tableau répétable : 1 ligne par matière ET par produit, à partir de la ligne 6
    "3. Composition des produits": [
        ("Référence produit", "A6"),
        ("Matière / composant", "B6"),
        ("Masse pour 1 unité de produit (kg)", "C6"),
        ("Matière recyclée ?", "D6"),
    ],

    # CONDITIONNEL : à remplir uniquement si déclaration GEREP / ICPE
    "6. Rejets air-eau (GEREP)": [
        # (optionnel — valeurs par défaut sinon)
    ],

    # INDICATIF : une estimation suffit (valeurs normalisées sinon)
    "8. Livraison et fin de vie": [
        ("Mode de transport principal vers vos clients", "C6"),
        ("Distance site → client (km)", "C7"),
        ("Durée de vie estimée du produit (années)", "C10"),
        ("Que devient le produit en fin de vie", "C14"),
    ]
}

# Liste blanche d'unites par onglet/colonne. Aucune dropdown detectee dans la
# trame, donc rien n'est presume. A remplir si les unites sont controlees.
# Ex. {"4. Energie - Eau": {"col": 4, "autorisees": {"kWh", "MWh", "L", "m3"}}}
UNITES_AUTORISEES = {
    "2. Vos produits":               {"col": 8, "autorisees": {"kg", "g", "t", "u", "m", "m2", "m3"}},
    "2. Vos produits":               {"col": 10, "autorisees": {"kg", "g", "t", "u", "m", "m2", "m3"}},
    "3. Composition des produits":   {"col": 3, "autorisees": {"kg", "g"}},
    "4. Énergie - Eau":              {"col": 4, "autorisees": {"kWh", "MWh", "L", "m3", "kg", "MJ"}},
    "5. Déchets":                    {"col": 4, "autorisees": {"kg", "t", "l", "m3", "u"}},
    "6. Rejets air-eau (GEREP)":     {"col": 4, "autorisees": {"kg", "g", "mg", "t", "tCO2", "m3"}},
    "7. Emballages":                 {"col": 4, "autorisees": {"kg", "g", "t", "u",}},
}

# Bornes physiques par champ pour le controle d'ordre de grandeur.
# Ex. {"4. Energie - Eau": {"col_valeur": 3, "min": 0, "max": 1e9}}
BORNES = {}


# ---------------------------------------------------------------------------
# OUTILS
# ---------------------------------------------------------------------------

def _norm(s):
    """Normalisation souple pour comparer les libelles malgre accents/casse/espaces."""
    if s is None:
        return ""
    import unicodedata
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def _vide(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


def _alerte(gravite, ou, description, action):
    return {"gravite": gravite, "ou": ou, "description": description, "action": action}


def _onglet_par_nom(wb, nom_attendu):
    """Retrouve un onglet par comparaison normalisee (tolere accents/casse)."""
    cible = _norm(nom_attendu)
    for ws in wb.worksheets:
        if _norm(ws.title) == cible:
            return ws
    return None


# ---------------------------------------------------------------------------
# CONTROLES
# ---------------------------------------------------------------------------

def controle_structure(wb, alertes):
    presents = {_norm(ws.title) for ws in wb.worksheets}
    for attendu in ONGLETS_ATTENDUS:
        if _norm(attendu) not in presents:
            gravite = "BLOQUANT" if attendu == ONGLET_BILAN else "BLOQUANT"
            alertes.append(_alerte(
                gravite, attendu, "onglet manquant ou renomme",
                "restaurer la trame de reference (/generer-trame-collecte)"))

    for nom, spec in ENTETES_ATTENDUS.items():
        ws = _onglet_par_nom(wb, nom)
        if ws is None:
            continue  # deja signale ci-dessus
        ligne, col0 = spec["ligne"], spec["col_depart"]
        reels = [_norm(ws.cell(ligne, col0 + i).value) for i in range(len(spec["libelles"]))]
        for i, attendu in enumerate(spec["libelles"]):
            cel = f"{openpyxl.utils.get_column_letter(col0 + i)}{ligne}"
            if _norm(attendu) not in reels:
                alertes.append(_alerte(
                    "BLOQUANT", f"{nom}!{cel}",
                    f"colonne attendue absente ou renommee : « {attendu} »",
                    "retablir l'en-tete d'origine"))


def controle_completude(wb, alertes):
    for nom, champs in CHAMPS_OBLIGATOIRES.items():
        ws = _onglet_par_nom(wb, nom)
        if ws is None:
            continue
        for libelle, cellule in champs:
            v = ws[cellule].value
            if _vide(v):
                alertes.append(_alerte(
                    "A_VERIFIER", f"{nom}!{cellule}",
                    f"champ obligatoire vide : {libelle}",
                    "completer ou confirmer non applicable avec le client"))


def controle_unites(wb, alertes):
    for nom, spec in UNITES_AUTORISEES.items():
        ws = _onglet_par_nom(wb, nom)
        if ws is None:
            continue
        col = spec["col"]
        autorisees = {_norm(u) for u in spec["autorisees"]}
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, col).value
            if _vide(v):
                continue
            if _norm(v) not in autorisees:
                cel = f"{openpyxl.utils.get_column_letter(col)}{r}"
                alertes.append(_alerte(
                    "A_VERIFIER", f"{nom}!{cel}",
                    f"unite inattendue : « {v} »",
                    "verifier l'unite avec le client (pas de conversion ici)"))


def controle_ordres_grandeur(wb, alertes):
    for nom, spec in BORNES.items():
        ws = _onglet_par_nom(wb, nom)
        if ws is None:
            continue
        col, mn, mx = spec["col_valeur"], spec.get("min"), spec.get("max")
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, col).value
            if not isinstance(v, (int, float)):
                continue
            cel = f"{openpyxl.utils.get_column_letter(col)}{r}"
            if mn is not None and v < mn:
                alertes.append(_alerte("A_VERIFIER", f"{nom}!{cel}",
                                       f"valeur sous la borne ({v} < {mn})",
                                       "verifier avec le client"))
            if mx is not None and v > mx:
                alertes.append(_alerte("A_VERIFIER", f"{nom}!{cel}",
                                       f"valeur au-dessus de la borne ({v} > {mx})",
                                       "verifier avec le client"))


def controle_bilan_massique(wb, alertes, tol):
    ws = _onglet_par_nom(wb, ONGLET_BILAN)
    if ws is None:
        return  # absence deja signalee en structure (BLOQUANT)

    totaux = {}
    for cle, rep in BILAN_REPERES.items():
        # Garde-fou : verifier le libelle avant de lire la valeur positionnelle.
        libelle = ws[rep["cellule_libelle"]].value
        if _norm(libelle) != _norm(cle):
            alertes.append(_alerte(
                "BLOQUANT", f"{ONGLET_BILAN}!{rep['cellule_libelle']}",
                f"repere « {cle} » introuvable a l'emplacement attendu "
                f"(lu : « {libelle} »)",
                "verifier que l'onglet bilan n'a pas ete modifie"))
            return
        val = ws[rep["cellule_valeur"]].value
        # Detection du cache de formule non recalcule.
        if val is None:
            alertes.append(_alerte(
                "BLOQUANT", f"{ONGLET_BILAN}!{rep['cellule_valeur']}",
                f"{cle} vide ou non recalcule "
                "(formule sans valeur en cache)",
                "ouvrir puis enregistrer le fichier dans Excel "
                "pour forcer le recalcul, puis relancer"))
            return
        if not isinstance(val, (int, float)):
            alertes.append(_alerte(
                "BLOQUANT", f"{ONGLET_BILAN}!{rep['cellule_valeur']}",
                f"{cle} non numerique (lu : « {val} »)",
                "verifier l'onglet bilan"))
            return
        totaux[cle] = float(val)

    entrants = totaux["TOTAL ENTRANTS"]
    sortants = totaux["TOTAL SORTANTS"]

    if entrants == 0:
        alertes.append(_alerte(
            "BLOQUANT", f"{ONGLET_BILAN}!C12",
            "TOTAL ENTRANTS = 0, ecart relatif non calculable",
            "collecter les masses entrantes (onglets 2, 3, 4, 7)"))
        return

    ecart_abs = sortants - entrants
    ecart_rel = ecart_abs / entrants
    if abs(ecart_rel) > tol:
        alertes.append(_alerte(
            "BLOQUANT", ONGLET_BILAN,
            f"bilan massique non boucle : ecart relatif "
            f"{ecart_rel * 100:.2f} % (> tolerance {tol * 100:.1f} %). "
            f"Entrants {entrants:.1f} kg/an, sortants {sortants:.1f} kg/an, "
            f"ecart {ecart_abs:+.1f} kg/an",
            "identifier le flux manquant ou en exces avec le client"))


# ---------------------------------------------------------------------------
# RENDU + LOG
# ---------------------------------------------------------------------------

_LABELS = {
    "BLOQUANT": ("\U0001F534", "Bloquant"),
    "A_VERIFIER": ("\U0001F7E0", "A verifier"),
    "INFO": ("\U0001F7E2", "Info"),
}
_ORDRE = ["BLOQUANT", "A_VERIFIER", "INFO"]


def rendre_markdown(alertes, fichier, tol):
    lignes = ["# Rapport de validation de la collecte", "",
              f"Fichier : `{os.path.basename(fichier)}`  ",
              f"Tolerance bilan massique : {tol * 100:.1f} %  ",
              f"Genere : {datetime.now(timezone.utc).isoformat(timespec='seconds')}", ""]
    if not alertes:
        lignes.append("Aucune alerte. Trame exploitable.")
        return "\n".join(lignes)
    par_grav = {g: [a for a in alertes if a["gravite"] == g] for g in _ORDRE}
    for g in _ORDRE:
        lot = par_grav[g]
        if not lot:
            continue
        emoji, titre = _LABELS[g]
        lignes.append(f"## {emoji} {titre} ({len(lot)})")
        for a in lot:
            lignes.append(f"- **{a['ou']}** : {a['description']}. _Action : {a['action']}._")
        lignes.append("")
    return "\n".join(lignes)


def ecrire_log(dossier_sortie, fichier, alertes, code, tol):
    log_dir = os.path.join(dossier_sortie, "_logs")
    os.makedirs(log_dir, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    chemin = os.path.join(log_dir, f"valider-fichier-collecte_{ts}.json")
    with open(chemin, "w", encoding="utf-8") as f:
        json.dump({
            "skill": "valider-fichier-collecte",
            "timestamp": ts,
            "fichier": os.path.basename(fichier),
            "tolerance_bilan": tol,
            "code_retour": code,
            "nb_alertes": len(alertes),
            "alertes": alertes,
        }, f, ensure_ascii=False, indent=2)
    return chemin


# ---------------------------------------------------------------------------
# ORCHESTRATION
# ---------------------------------------------------------------------------

def valider(fichier, dossier_sortie, tol=0.02):
    # data_only=True : on lit les valeurs en cache des formules. La detection
    # du cache absent (None) est geree dans le controle bilan.
    wb = openpyxl.load_workbook(fichier, data_only=True)
    alertes = []
    controle_structure(wb, alertes)
    controle_completude(wb, alertes)
    controle_unites(wb, alertes)
    controle_ordres_grandeur(wb, alertes)
    controle_bilan_massique(wb, alertes, tol)

    code = "BLOQUANT" if any(a["gravite"] == "BLOQUANT" for a in alertes) else "OK"

    os.makedirs(dossier_sortie, exist_ok=True)
    md = rendre_markdown(alertes, fichier, tol)
    chemin_md = os.path.join(dossier_sortie, "alertes-validation.md")
    with open(chemin_md, "w", encoding="utf-8") as f:
        f.write(md)
    chemin_log = ecrire_log(dossier_sortie, fichier, alertes, code, tol)
    return code, chemin_md, chemin_log


def main():
    p = argparse.ArgumentParser(description="Validation deterministe d'une trame de collecte FDES.")
    p.add_argument("fichier", help="chemin de la trame .xlsx remplie")
    p.add_argument("dossier_sortie", nargs="?", default=".",
                   help="dossier ou ecrire le rapport (defaut : courant)")
    p.add_argument("--tol", type=float, default=0.02,
                   help="tolerance du bilan massique (defaut 0.02 = 2%%)")
    args = p.parse_args()

    if not os.path.isfile(args.fichier):
        print(f"Erreur : fichier introuvable : {args.fichier}", file=sys.stderr)
        sys.exit(1)
    try:
        code, chemin_md, chemin_log = valider(args.fichier, args.dossier_sortie, args.tol)
    except Exception as e:  # erreur technique = ni OK ni BLOQUANT metier
        print(f"Erreur technique : {e}", file=sys.stderr)
        sys.exit(1)

    print(f"Code retour : {code}")
    print(f"Rapport     : {chemin_md}")
    print(f"Log         : {chemin_log}")
    sys.exit(0 if code == "OK" else 2)


if __name__ == "__main__":
    main()
