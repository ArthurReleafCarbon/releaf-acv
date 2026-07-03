"""
emit_contrat.py - Émission du contrat de structure (log de génération).

`generer-trame-collecte` écrit, en fin de traitement, un log Markdown dont le
frontmatter YAML décrit la structure RÉELLEMENT appliquée à la trame. Ce contrat est
consommé par `verifier-trame-collecte` (voir
../../verifier-trame-collecte/reference/contrat-structure.md).

Principe : dériver le maximum du classeur généré (onglets présents, colonnes d'attributs
de l'onglet 2), pour GARANTIR que le contrat colle à la structure réelle et non aux
intentions du script pilote. Le pilote ne fournit que ce qui n'est pas dérivable :
la politique produit (postes d'énergie obligatoires, listes blanches d'unités,
valeurs fermées, bornes).

Usage dans le script pilote :

    from emit_contrat import deriver_contrat, ecrire_log
    contrat = deriver_contrat(
        wb,
        scenario={"produits": "mono", "sites": "mono"},
        produit={"type": "isolant PSE", "procede": "extrusion"},
        overrides={
            "postes_energie_obligatoires": ["Électricité du réseau"],
            "unites_attendues": {"5. Déchets": {"C": ["kg", "tonnes", "t"]}},
            "valeurs_fermees": {"3. Composition des produits": {"D": ["Oui", "Non", "Inconnu"]}},
            "bornes": {"2. Vos produits": {"G": {"min": 0, "max": 100000}}},
        },
    )
    ecrire_log(chemin_log, contrat, texte_humain="Diff appliqué : ...")
"""
from openpyxl.utils import get_column_letter, column_index_from_string

try:
    import yaml
except ImportError:  # pragma: no cover
    yaml = None

SCHEMA_VERSION = 1

# Onglet d'attributs produit et libelles-reperes stables.
ONGLET_PRODUITS = "2. Vos produits"
LIB_REFERENCE = "Référence interne / code"
LIB_MASSE = "Masse d'une unité (kg)"

# Champs obligatoires stables (par convention de la trame, cf. base_layout.md).
# Le script pilote peut les surcharger via overrides["champs_obligatoires"].
CHAMPS_OBLIGATOIRES_DEFAUT = {
    "2. Vos produits": ["A", "G", "H", "I", "J"],
    "3. Composition des produits": ["A", "B", "C"],
}


def _norme(v):
    return " ".join(str(v).split()).lower() if v is not None else ""


def _trouver_ligne_entete(ws, libelle, col=1, max_row=15):
    cible = _norme(libelle)
    for r in range(1, min(max_row, ws.max_row) + 1):
        if _norme(ws.cell(row=r, column=col).value) == cible:
            return r
    return None


def colonnes_attributs_produit(ws):
    """Dérive les colonnes d'attributs de l'onglet 2 (celles entre Référence/Description
    et Masse). Ce sont les colonnes qui VARIENT selon le produit (D, E, F...)."""
    hr = _trouver_ligne_entete(ws, LIB_REFERENCE, col=1) or 4
    # colonne de la masse = borne droite des attributs
    col_masse = None
    for c in range(1, ws.max_column + 1):
        if _norme(ws.cell(row=hr, column=c).value) == _norme(LIB_MASSE):
            col_masse = c
            break
    if col_masse is None:
        return []
    # attributs = colonnes entre D (4) et col_masse - 1 inclus, non vides
    out = []
    for c in range(column_index_from_string("D"), col_masse):
        libelle = ws.cell(row=hr, column=c).value
        if libelle is not None and str(libelle).strip():
            out.append({"col": get_column_letter(c), "libelle": str(libelle).strip(),
                        "obligatoire": False})
    return out


def deriver_contrat(wb, scenario, produit, overrides=None):
    """Construit le dict contrat en dérivant du classeur ce qui est dérivable,
    puis en fusionnant les overrides fournis par le script pilote."""
    overrides = overrides or {}
    contrat = {
        "schema_version": SCHEMA_VERSION,
        "scenario": dict(scenario),
        "produit": dict(produit),
        "onglets_presents": list(wb.sheetnames),
        "onglets_retires": overrides.get("onglets_retires", []),
        "colonnes_produit": overrides.get("colonnes_produit")
        or (colonnes_attributs_produit(wb[ONGLET_PRODUITS])
            if ONGLET_PRODUITS in wb.sheetnames else []),
        "champs_obligatoires": overrides.get(
            "champs_obligatoires",
            # ne garder que les onglets reellement presents
            {k: v for k, v in CHAMPS_OBLIGATOIRES_DEFAUT.items() if k in wb.sheetnames}),
    }
    # champs facultatifs : uniquement s'ils sont fournis (sinon controle saute cote validateur)
    for cle in ("postes_energie_obligatoires", "unites_attendues", "valeurs_fermees", "bornes"):
        if cle in overrides:
            contrat[cle] = overrides[cle]
    return contrat


def _dump_yaml(contrat):
    if yaml is not None:
        return yaml.safe_dump(contrat, allow_unicode=True, sort_keys=False,
                              default_flow_style=False).rstrip()
    # repli minimal sans PyYAML (non recommande : PyYAML est requis par le validateur)
    raise RuntimeError("PyYAML requis pour émettre le contrat (pip install pyyaml).")


def ecrire_log(chemin, contrat, texte_humain=""):
    """Écrit le log de génération : frontmatter YAML (contrat) + texte libre pour l'humain."""
    corps = _dump_yaml(contrat)
    contenu = f"---\n{corps}\n---\n\n{texte_humain}\n"
    with open(chemin, "w", encoding="utf-8") as f:
        f.write(contenu)
    return chemin
