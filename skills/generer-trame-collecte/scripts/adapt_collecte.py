"""
Primitives d'adaptation STRUCTURELLE de la trame de collecte FDES.

Regle d'or : ce module ne modifie QUE des champs (en-tetes, libelles, listes
deroulantes, lignes/colonnes/onglets pertinents). Il n'ecrit JAMAIS de valeur
de donnee ni de nomenclature produit. Voir SKILL.md, section "Structure vs contenu".

Usage : importer ces fonctions dans un court script pilote ecrit par Claude
pour le produit en cours, puis recalculer avec scripts/recalc.py du skill xlsx.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---- Palette exacte de la base (ne pas changer) ----
FONT  = "Arial"
C_HDR = "2E5E3A"   # bandeau section
C_COL = "D6E4D0"   # entete colonne
C_IN  = "FFF6CC"   # cellule a remplir (jaune)
C_LOCK= "ECECEC"   # cellule verrouillee / BE
C_NOTE= "F5F5F5"   # note
C_YEAR= "FFF2CC"   # bandeau rappel annee
WHITE = "FFFFFF"

_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _font(b=False, c="000000", sz=10, it=False):
    return Font(name=FONT, bold=b, color=c, size=sz, italic=it)

def _fill(hex_):
    return PatternFill("solid", fgColor=hex_)

# ---------- I/O ----------
def open_base(path):
    return load_workbook(path)

def save(wb, path):
    wb.save(path)

# ---------- Styles de cellule ----------
def style_input(ws, row, col):
    c = ws.cell(row=row, column=col)
    c.fill = _fill(C_IN); c.border = BORDER; c.font = _font()
    c.alignment = Alignment(horizontal="left", vertical="center")
    return c

def style_label(ws, row, col, text=None):
    c = ws.cell(row=row, column=col)
    if text is not None:
        c.value = text
    c.fill = _fill(C_LOCK); c.border = BORDER; c.font = _font()
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    return c

def style_colhead(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(C_COL); c.border = BORDER; c.font = _font(b=True, sz=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

# ---------- Operations STRUCTURE ----------
def set_colheaders(ws, row, headers, start_col=1):
    """Reecrit une ligne d'en-tete de colonnes. headers = liste de str."""
    for i, h in enumerate(headers):
        style_colhead(ws, row, start_col + i, h)

def set_header_text(ws, row, col, text):
    """Change un libelle (en-tete colonne ou libelle de ligne) sans toucher le style."""
    ws.cell(row=row, column=col).value = text

def add_input_column(ws, header_row, col, header, first_row, last_row):
    """Ajoute une colonne d'attribut : en-tete + cellules jaunes vides."""
    style_colhead(ws, header_row, col, header)
    for r in range(first_row, last_row + 1):
        style_input(ws, r, col)

def add_labelled_rows(ws, start_row, labels, label_col, input_cols):
    """Ajoute des lignes a libelle (champ vide a remplir). NE PRE-REMPLIT PAS de valeur.
    labels = liste de libelles ; input_cols = liste d'indices de colonnes a rendre jaunes."""
    r = start_row
    for lab in labels:
        style_label(ws, r, label_col, lab)
        for c in input_cols:
            style_input(ws, r, c)
        r += 1
    return r

def delete_rows(ws, first_row, count=1):
    ws.delete_rows(first_row, count)

def insert_rows(ws, before_row, count=1):
    ws.insert_rows(before_row, count)

# ---------- Aucune liste deroulante ----------
# Decision Thomas : la trame ne contient AUCUNE liste deroulante (validation de
# donnees). Les unites et valeurs attendues sont indiquees en clair dans les
# libelles (cf. set_header_text). strip_dropdowns purge toute validation heritee.
def strip_dropdowns(ws):
    """Filet de securite : retire toute validation de donnees de la feuille."""
    ws.data_validations.dataValidation = []

def strip_all_dropdowns(wb):
    for ws in wb.worksheets:
        strip_dropdowns(ws)

# ---------- Onglets ----------
def delete_sheet(wb, name):
    if name in wb.sheetnames:
        del wb[name]

def hide_sheet(wb, name):
    if name in wb.sheetnames:
        wb[name].sheet_state = "hidden"

def rename_sheet(wb, old, new):
    if old in wb.sheetnames:
        wb[old].title = new
