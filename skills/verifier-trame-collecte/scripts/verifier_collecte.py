"""
verifier_collecte.py - Validation deterministe d'une trame de collecte FDES remplie.

Couche C1 (Python pur). La skill SIGNALE, elle ne CORRIGE jamais. Elle lit la structure
attendue dans le frontmatter YAML du log de generation (voir reference/contrat-structure.md),
la compare a la trame remplie par le client, et produit un rapport d'alertes classe par gravite.

Usage :
    python verifier_collecte.py <trame.xlsx> <log_generation.md> <sortie_dir> \
        [--tol 0.02] [--seuil-dur 0.10]

Codes retour :
    0 = OK (exploitable, eventuelles alertes non bloquantes)
    2 = BLOQUANT (au moins une alerte rouge)
    1 = erreur technique (fichier illisible, contrat absent/invalide)
"""
import argparse
import datetime as _dt
import os
import sys
import unicodedata

try:
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string
except ImportError:
    sys.stderr.write("ERREUR TECHNIQUE : openpyxl requis.\n")
    sys.exit(1)

try:
    import yaml
except ImportError:
    sys.stderr.write("ERREUR TECHNIQUE : PyYAML requis (pip install pyyaml).\n")
    sys.exit(1)

# ------------------------------------------------------------------ gravites
ROUGE = "rouge"
ORANGE = "orange"
VERT = "vert"
_LABEL = {ROUGE: "🔴 Bloquant", ORANGE: "🟠 À vérifier", VERT: "🟢 Info"}
_ORDER = [ROUGE, ORANGE, VERT]


class Alerte:
    def __init__(self, gravite, onglet, cellule, description, action):
        self.gravite = gravite
        self.onglet = onglet
        self.cellule = cellule
        self.description = description
        self.action = action


# ---------------------------------------------------------------- invariants
# Ancres fixes du gabarit (voir generer-trame-collecte/reference/base_layout.md).
# Ne portent QUE ce qui ne varie pas selon le produit.
ONGLET_PRODUITS = "2. Vos produits"
ONGLET_COMPO = "3. Composition des produits"
ONGLET_ENERGIE = "4. Consommations usines"

# Derniere ligne de donnees exploitable par onglet (borne les scans).
# Onglet 3 : bloc 1 nomenclature 6->46 (le bloc 2 "produits consommes" a une autre
# structure et n'est pas controle en v1). None = jusqu'a max_row.
DERNIERE_LIGNE = {
    ONGLET_PRODUITS: 54,
    ONGLET_COMPO: 46,
}

# Onglet 2 : en-tetes de colonnes STABLES (col_lettre -> libelle attendu)
ENTETES_FIXES_PRODUITS = {
    "A": "Référence interne / code",
    "G": "Masse d'une unité (kg)",
    "H": "Unité de vente",
    "I": "Quantité fabriquée sur l'année",
    "J": "Unité de cette quantité",
}


# ------------------------------------------------------------------ utilitaires
def norme(v):
    """Normalise pour comparaison robuste : str, sans accents, minuscule, espaces reduits."""
    if v is None:
        return ""
    s = unicodedata.normalize("NFKD", str(v))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def cell_text(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return v


def est_vide(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


def to_float(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        t = v.strip().replace(",", ".").replace(" ", "")
        try:
            return float(t)
        except ValueError:
            return None
    return None


def find_header_row(ws, label_attendu, col=1, defaut=None, max_row=15):
    """Localise la ligne d'en-tete en cherchant un libelle connu (robuste au decalage de lignes)."""
    cible = norme(label_attendu)
    for r in range(1, min(max_row, ws.max_row) + 1):
        if norme(cell_text(ws, r, col)) == cible:
            return r
    return defaut


def find_row_by_label(ws, label, col=1, max_row=None):
    cible = norme(label)
    top = max_row or ws.max_row
    for r in range(1, top + 1):
        if norme(cell_text(ws, r, col)) == cible:
            return r
    return None


# ------------------------------------------------------------------ contrat
def charger_contrat(chemin_log):
    """Extrait et parse le frontmatter YAML du log de generation."""
    if not os.path.isfile(chemin_log):
        raise SystemExit1(f"log de generation introuvable : {chemin_log}")
    with open(chemin_log, "r", encoding="utf-8") as f:
        contenu = f.read()
    if not contenu.lstrip().startswith("---"):
        raise SystemExit1("le log ne commence pas par un frontmatter YAML (--- ... ---).")
    # premier bloc delimite par --- ... ---
    apres = contenu.split("---", 2)
    if len(apres) < 3:
        raise SystemExit1("frontmatter YAML incomplet dans le log.")
    try:
        data = yaml.safe_load(apres[1])
    except yaml.YAMLError as e:
        raise SystemExit1(f"frontmatter YAML illisible : {e}")
    if not isinstance(data, dict):
        raise SystemExit1("frontmatter YAML vide ou mal forme.")
    if "schema_version" not in data or "onglets_presents" not in data:
        raise SystemExit1("contrat incomplet : 'schema_version' et 'onglets_presents' requis.")
    return data


class SystemExit1(Exception):
    """Erreur technique -> code retour 1."""


# ------------------------------------------------------------------ familles
def famille1_structure(wb, contrat, al):
    attendus = contrat.get("onglets_presents", [])
    reels = set(norme(t) for t in wb.sheetnames)
    for t in attendus:
        if norme(t) not in reels:
            al.append(Alerte(ROUGE, t, "-", "Onglet attendu manquant dans la trame reçue.",
                             "Vérifier que le client a bien renvoyé le fichier généré, non modifié."))
    attendus_norm = set(norme(t) for t in attendus)
    for t in wb.sheetnames:
        if norme(t) not in attendus_norm:
            al.append(Alerte(ORANGE, t, "-", "Onglet non prévu par le contrat de génération.",
                             "Le client a peut-être ajouté un onglet ; vérifier son contenu."))
    # noms definis
    noms = set(wb.defined_names.keys()) if hasattr(wb.defined_names, "keys") else set(
        n.name for n in wb.defined_names.definedName)
    for nd in ("AnneeRef", "ListeProduits"):
        if nd not in noms:
            al.append(Alerte(ROUGE, "-", nd, f"Nom défini « {nd} » absent.",
                             "Structure de la trame altérée ; redemander le fichier d'origine."))
    # onglet 2 : en-tetes fixes + colonnes produit adaptees
    if ONGLET_PRODUITS in wb.sheetnames:
        ws = wb[ONGLET_PRODUITS]
        hr = find_header_row(ws, ENTETES_FIXES_PRODUITS["A"], col=1, defaut=4)
        for lettre, libelle in ENTETES_FIXES_PRODUITS.items():
            ci = column_index_from_string(lettre)
            if norme(cell_text(ws, hr, ci)) != norme(libelle):
                al.append(Alerte(ROUGE, ONGLET_PRODUITS, f"{lettre}{hr}",
                                 f"En-tête fixe renommé ou absent (attendu « {libelle} »).",
                                 "Ne pas renommer les colonnes ; repartir de la trame d'origine."))
        for c in contrat.get("colonnes_produit", []):
            lettre = c.get("col")
            libelle = c.get("libelle", "")
            ci = column_index_from_string(lettre)
            if norme(cell_text(ws, hr, ci)) != norme(libelle):
                al.append(Alerte(ROUGE, ONGLET_PRODUITS, f"{lettre}{hr}",
                                 f"Colonne d'attribut produit modifiée (attendu « {libelle} »).",
                                 "Colonne adaptée au produit altérée par le client."))


def _lignes_actives(ws, hr, col_cle, dernier=None):
    """Lignes de saisie 'actives' = celles dont la colonne cle (souvent A) est non vide."""
    top = dernier or ws.max_row
    ci = column_index_from_string(col_cle) if isinstance(col_cle, str) else col_cle
    out = []
    for r in range(hr + 1, top + 1):
        if not est_vide(cell_text(ws, r, ci)):
            out.append(r)
    return out


def famille2_completude(wb, contrat, al):
    # annee de reference (AnneeRef -> B16 par defaut, localise si decale)
    if "1. Société et site" in wb.sheetnames:
        ws1 = wb["1. Société et site"]
        r = find_row_by_label(ws1, "Année de référence des données", col=1) or 16
        if est_vide(cell_text(ws1, r, 2)):
            al.append(Alerte(ROUGE, "1. Société et site", f"B{r}",
                             "Année de référence non renseignée.",
                             "Sans année de référence, aucune répartition possible."))
    oblig = contrat.get("champs_obligatoires", {})
    for onglet, cols in oblig.items():
        if onglet not in wb.sheetnames:
            continue
        ws = wb[onglet]
        # localiser la ligne d'en-tete par la 1re colonne obligatoire connue si possible
        hr = 4 if onglet == ONGLET_PRODUITS else (5 if onglet == ONGLET_COMPO else 4)
        col_cle = cols[0] if cols else "A"
        for r in _lignes_actives(ws, hr, col_cle, dernier=DERNIERE_LIGNE.get(onglet)):
            for lettre in cols:
                ci = column_index_from_string(lettre)
                if est_vide(cell_text(ws, r, ci)):
                    al.append(Alerte(ROUGE, onglet, f"{lettre}{r}",
                                     "Champ obligatoire vide sur une ligne renseignée.",
                                     "Compléter ou supprimer la ligne si non pertinente."))
    # sources (traçabilite) sur composition -> orange
    if ONGLET_COMPO in wb.sheetnames:
        ws = wb[ONGLET_COMPO]
        hr = find_header_row(ws, "Référence produit", col=1, defaut=5)
        h_source = column_index_from_string("H")
        for r in _lignes_actives(ws, hr, "A", dernier=46):
            if est_vide(cell_text(ws, r, h_source)):
                al.append(Alerte(ORANGE, ONGLET_COMPO, f"H{r}",
                                 "Source de l'information non renseignée.",
                                 "À lever avec le client (traçabilité)."))
    # postes energie obligatoires
    postes = contrat.get("postes_energie_obligatoires", [])
    if postes and ONGLET_ENERGIE in wb.sheetnames:
        ws = wb[ONGLET_ENERGIE]
        col_val = column_index_from_string("B")
        for lab in postes:
            r = find_row_by_label(ws, lab, col=1)
            if r is None:
                al.append(Alerte(ORANGE, ONGLET_ENERGIE, "-",
                                 f"Poste d'énergie obligatoire « {lab} » introuvable.",
                                 "Vérifier que la ligne n'a pas été supprimée."))
            elif est_vide(cell_text(ws, r, col_val)):
                al.append(Alerte(ROUGE, ONGLET_ENERGIE, f"B{r}",
                                 f"Poste d'énergie obligatoire « {lab} » non renseigné.",
                                 "Consommation indispensable au calcul pour ce procédé."))


def _refs_produits(wb):
    """Ensemble des references produit declarees onglet 2 (col A, lignes actives)."""
    refs = {}
    if ONGLET_PRODUITS not in wb.sheetnames:
        return refs
    ws = wb[ONGLET_PRODUITS]
    hr = find_header_row(ws, ENTETES_FIXES_PRODUITS["A"], col=1, defaut=4)
    for r in _lignes_actives(ws, hr, "A", dernier=54):
        ref = cell_text(ws, r, 1)
        refs[norme(ref)] = (str(ref).strip(), r)
    return refs


def famille3_coherence(wb, contrat, al):
    refs = _refs_produits(wb)
    # references orphelines + produits sans composition
    refs_utilisees = set()
    if ONGLET_COMPO in wb.sheetnames:
        ws = wb[ONGLET_COMPO]
        hr = find_header_row(ws, "Référence produit", col=1, defaut=5)
        for r in _lignes_actives(ws, hr, "A", dernier=46):
            ref = norme(cell_text(ws, r, 1))
            refs_utilisees.add(ref)
            if ref not in refs:
                al.append(Alerte(ROUGE, ONGLET_COMPO, f"A{r}",
                                 f"Référence produit « {cell_text(ws, r, 1)} » absente de l'onglet 2.",
                                 "Corriger la référence ou la déclarer dans « Vos produits »."))
    for ref_n, (ref_aff, r) in refs.items():
        if ref_n not in refs_utilisees:
            al.append(Alerte(ROUGE, ONGLET_PRODUITS, f"A{r}",
                             f"Produit « {ref_aff} » sans aucune ligne de composition.",
                             "Renseigner sa nomenclature dans l'onglet 3."))
    # unites attendues
    for onglet, cols in contrat.get("unites_attendues", {}).items():
        if onglet not in wb.sheetnames:
            continue
        ws = wb[onglet]
        hr = 5 if onglet == ONGLET_COMPO else 4
        for lettre, blanche in cols.items():
            ci = column_index_from_string(lettre)
            blanche_n = set(norme(u) for u in blanche)
            for r in _lignes_actives(ws, hr, "A", dernier=DERNIERE_LIGNE.get(onglet)):
                v = cell_text(ws, r, ci)
                if not est_vide(v) and norme(v) not in blanche_n:
                    al.append(Alerte(ORANGE, onglet, f"{lettre}{r}",
                                     f"Unité « {v} » hors liste attendue ({', '.join(blanche)}).",
                                     "Convertir en amont (convertir-unites) ou vérifier avec le client."))
    # valeurs fermees
    for onglet, cols in contrat.get("valeurs_fermees", {}).items():
        if onglet not in wb.sheetnames:
            continue
        ws = wb[onglet]
        hr = 5 if onglet == ONGLET_COMPO else 4
        for lettre, autor in cols.items():
            ci = column_index_from_string(lettre)
            autor_n = set(norme(u) for u in autor)
            for r in _lignes_actives(ws, hr, "A", dernier=DERNIERE_LIGNE.get(onglet)):
                v = cell_text(ws, r, ci)
                if not est_vide(v) and norme(v) not in autor_n:
                    al.append(Alerte(ORANGE, onglet, f"{lettre}{r}",
                                     f"Valeur « {v} » non autorisée ({', '.join(autor)}).",
                                     "Utiliser une des valeurs prévues."))


def famille4_bilan_masse(wb, contrat, al, tol, seuil_dur):
    if ONGLET_PRODUITS not in wb.sheetnames or ONGLET_COMPO not in wb.sheetnames:
        return
    wsp = wb[ONGLET_PRODUITS]
    wsc = wb[ONGLET_COMPO]
    hrp = find_header_row(wsp, ENTETES_FIXES_PRODUITS["A"], col=1, defaut=4)
    hrc = find_header_row(wsc, "Référence produit", col=1, defaut=5)
    col_masse_prod = column_index_from_string("G")
    col_ref_compo = column_index_from_string("A")
    col_masse_compo = column_index_from_string("C")
    # somme composants par reference
    sommes = {}
    for r in _lignes_actives(wsc, hrc, "A", dernier=46):
        ref = norme(cell_text(wsc, r, col_ref_compo))
        m = to_float(cell_text(wsc, r, col_masse_compo))
        if m is not None:
            sommes[ref] = sommes.get(ref, 0.0) + m
    for r in _lignes_actives(wsp, hrp, "A", dernier=54):
        ref = norme(cell_text(wsp, r, 1))
        masse = to_float(cell_text(wsp, r, col_masse_prod))
        if masse is None or masse <= 0:
            continue  # deja signale ailleurs
        if ref not in sommes:
            continue  # produit sans compo deja signale F3
        ecart = abs(sommes[ref] - masse) / masse
        aff = cell_text(wsp, r, 1)
        if ecart <= tol:
            continue
        elif ecart <= seuil_dur:
            al.append(Alerte(ORANGE, ONGLET_COMPO, f"C (produit {aff})",
                             f"Bilan massique produit : écart {ecart*100:.1f}% "
                             f"(somme composants {sommes[ref]:.4g} kg vs masse unité {masse:.4g} kg).",
                             "À lever avec le client : nomenclature peut-être incomplète."))
        else:
            al.append(Alerte(ROUGE, ONGLET_COMPO, f"C (produit {aff})",
                             f"Bilan massique produit FAUX : écart {ecart*100:.1f}% "
                             f"(somme composants {sommes[ref]:.4g} kg vs masse unité {masse:.4g} kg).",
                             "Écart trop important ; corriger avant import."))


def famille5_plausibilite(wb, contrat, al):
    # masses/quantites negatives sur les onglets tabulaires connus
    cibles = {
        ONGLET_PRODUITS: (4, ["G", "I", "K"]),
        ONGLET_COMPO: (5, ["C"]),
        ONGLET_ENERGIE: (4, ["B"]),
        "5. Déchets": (4, ["B"]),
    }
    for onglet, (hr_def, lettres) in cibles.items():
        if onglet not in wb.sheetnames:
            continue
        ws = wb[onglet]
        for lettre in lettres:
            ci = column_index_from_string(lettre)
            for r in range(hr_def + 1, ws.max_row + 1):
                v = to_float(cell_text(ws, r, ci))
                if v is not None and v < 0:
                    al.append(Alerte(ROUGE, onglet, f"{lettre}{r}",
                                     f"Valeur négative ({v}).",
                                     "Une masse ou quantité ne peut pas être négative."))
    # masse unite produit nulle
    if ONGLET_PRODUITS in wb.sheetnames:
        ws = wb[ONGLET_PRODUITS]
        hr = find_header_row(ws, ENTETES_FIXES_PRODUITS["A"], col=1, defaut=4)
        cg = column_index_from_string("G")
        for r in _lignes_actives(ws, hr, "A", dernier=54):
            v = to_float(cell_text(ws, r, cg))
            if v is not None and v == 0:
                al.append(Alerte(ROUGE, ONGLET_PRODUITS, f"G{r}",
                                 "Masse d'une unité nulle.",
                                 "Bloque le calcul ; renseigner une masse non nulle."))
    # bornes du contrat
    for onglet, cols in contrat.get("bornes", {}).items():
        if onglet not in wb.sheetnames:
            continue
        ws = wb[onglet]
        hr = 4 if onglet != ONGLET_COMPO else 5
        for lettre, b in cols.items():
            ci = column_index_from_string(lettre)
            mn, mx = b.get("min"), b.get("max")
            for r in _lignes_actives(ws, hr, "A"):
                v = to_float(cell_text(ws, r, ci))
                if v is None:
                    continue
                if (mn is not None and v < mn) or (mx is not None and v > mx):
                    al.append(Alerte(ORANGE, onglet, f"{lettre}{r}",
                                     f"Valeur {v} hors bornes plausibles [{mn}, {mx}].",
                                     "Vérifier l'ordre de grandeur avec le client."))


# ------------------------------------------------------------------ rapport
def ecrire_rapport(al, contrat, chemin, tol, seuil_dur, sautes):
    par = {g: [a for a in al if a.gravite == g] for g in _ORDER}
    lignes = []
    lignes.append("# Rapport de validation de la trame de collecte\n")
    prod = contrat.get("produit", {})
    lignes.append(f"- Produit : {prod.get('type', 'n.c.')} — procédé : {prod.get('procede', 'n.c.')}")
    sc = contrat.get("scenario", {})
    lignes.append(f"- Scénario : {sc.get('produits', '?')} produit / {sc.get('sites', '?')} site")
    lignes.append(f"- Tolérance bilan massique : {tol*100:.1f}% (seuil bloquant {seuil_dur*100:.1f}%)")
    statut = "BLOQUANT" if par[ROUGE] else "OK"
    lignes.append(f"- **Statut : {statut}** — "
                  f"{len(par[ROUGE])} 🔴, {len(par[ORANGE])} 🟠, {len(par[VERT])} 🟢\n")
    for g in _ORDER:
        if not par[g] and g != VERT:
            continue
        lignes.append(f"\n## {_LABEL[g]}\n")
        if not par[g]:
            lignes.append("_Aucune alerte._")
        for a in par[g]:
            lignes.append(f"- **{a.onglet} · {a.cellule}** — {a.description} "
                          f"_Action : {a.action}_")
    if sautes:
        lignes.append("\n## 🟢 Contrôles non effectués (contrat incomplet)\n")
        for s in sautes:
            lignes.append(f"- {s}")
    lignes.append("\n---\n_Validation déterministe. La skill signale, elle ne corrige pas. "
                  "Un 🟠 n'est pas un blocage : c'est au praticien de décider._")
    with open(chemin, "w", encoding="utf-8") as f:
        f.write("\n".join(lignes))


def ecrire_log(sortie_dir, args, al, code):
    logs = os.path.join(sortie_dir, "_logs")
    os.makedirs(logs, exist_ok=True)
    ts = _dt.datetime.now().strftime("%Y%m%dT%H%M%S")
    n = {g: len([a for a in al if a.gravite == g]) for g in _ORDER}
    with open(os.path.join(logs, f"verifier-trame-collecte_{ts}.md"), "w", encoding="utf-8") as f:
        f.write(f"""---
skill: verifier-trame-collecte
horodatage: {ts}
trame: {os.path.basename(args.trame)}
log_generation: {os.path.basename(args.log)}
tol: {args.tol}
seuil_dur: {args.seuil_dur}
code_retour: {code}
alertes: {{rouge: {n[ROUGE]}, orange: {n[ORANGE]}, vert: {n[VERT]}}}
---

Validation exécutée le {ts}. Statut : {'BLOQUANT' if code == 2 else 'OK'}.
""")


# ------------------------------------------------------------------ main
def main():
    ap = argparse.ArgumentParser(description="Validation d'une trame de collecte FDES remplie.")
    ap.add_argument("trame", help="trame remplie (.xlsx)")
    ap.add_argument("log", help="log de génération (.md, frontmatter YAML)")
    ap.add_argument("sortie", help="dossier de sortie")
    ap.add_argument("--tol", type=float, default=0.02, help="tolérance bilan massique (défaut 0.02)")
    ap.add_argument("--seuil-dur", dest="seuil_dur", type=float, default=0.10,
                    help="seuil bloquant bilan massique (défaut 0.10)")
    args = ap.parse_args()

    try:
        contrat = charger_contrat(args.log)
    except SystemExit1 as e:
        sys.stderr.write(f"ERREUR TECHNIQUE : {e}\n")
        sys.exit(1)
    if not os.path.isfile(args.trame):
        sys.stderr.write(f"ERREUR TECHNIQUE : trame introuvable : {args.trame}\n")
        sys.exit(1)
    try:
        wb = load_workbook(args.trame, data_only=True)
    except Exception as e:
        sys.stderr.write(f"ERREUR TECHNIQUE : trame illisible : {e}\n")
        sys.exit(1)

    sautes = []
    if "unites_attendues" not in contrat:
        sautes.append("Cohérence des unités (pas de liste blanche dans le contrat).")
    if "bornes" not in contrat:
        sautes.append("Ordres de grandeur (pas de bornes dans le contrat).")

    al = []
    famille1_structure(wb, contrat, al)
    famille2_completude(wb, contrat, al)
    famille3_coherence(wb, contrat, al)
    famille4_bilan_masse(wb, contrat, al, args.tol, args.seuil_dur)
    famille5_plausibilite(wb, contrat, al)

    os.makedirs(args.sortie, exist_ok=True)
    rapport = os.path.join(args.sortie, "alertes-validation.md")
    ecrire_rapport(al, contrat, rapport, args.tol, args.seuil_dur, sautes)

    code = 2 if any(a.gravite == ROUGE for a in al) else 0
    ecrire_log(args.sortie, args, al, code)
    n = {g: len([a for a in al if a.gravite == g]) for g in _ORDER}
    sys.stdout.write(f"Statut : {'BLOQUANT' if code == 2 else 'OK'} | "
                     f"{n[ROUGE]} 🔴, {n[ORANGE]} 🟠, {n[VERT]} 🟢\n")
    sys.stdout.write(f"Rapport : {rapport}\n")
    sys.exit(code)


if __name__ == "__main__":
    main()

# fin du module
