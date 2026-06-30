# -*- coding: utf-8 -*-
"""
Remplir_FDES — outil de remplissage automatique des FDES.
Double-cliquez ce fichier : il vous demande le modèle Word et le
tableur Excel, puis génère la FDES complétée à côté du modèle.
"""
import sys, subprocess, os

# --- installe automatiquement les bibliotheques manquantes ---
def _ensure(pkg, import_name=None):
    try:
        __import__(import_name or pkg)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])

_ensure("python-docx", "docx")
_ensure("openpyxl")

from docx import Document
import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox

W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
XMLSPACE = "{http://www.w3.org/XML/1998/namespace}space"


def lire_valeurs(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    # cherche une feuille avec une colonne Tag ; sinon prend l'active
    feuilles = [wb["Contrôles FDES"]] if "Contrôles FDES" in wb.sheetnames else list(wb.worksheets)
    for ws in feuilles + list(wb.worksheets):
        headers = {}
        for j, cell in enumerate(ws[1]):
            if cell.value:
                headers[str(cell.value).strip().lower()] = j
        col_tag = next((j for k, j in headers.items() if k.startswith("tag")), None)
        col_val = next((j for k, j in headers.items() if k.startswith("valeur")), None)
        if col_tag is None or col_val is None:
            continue
        vals = {}
        for row in ws.iter_rows(min_row=2):
            tag = row[col_tag].value
            val = row[col_val].value
            if tag is not None and val is not None and str(val).strip() != "":
                vals[str(tag).strip()] = str(val)
        if vals:
            return vals
    return {}


def remplir_sdt(sdt, valeur):
    content = sdt.find(W + "sdtContent")
    if content is None:
        return False
    ts = list(content.iter(W + "t"))
    if ts:
        ts[0].text = valeur
        for extra in ts[1:]:
            extra.text = ""
        return True
    r = content.find(".//" + W + "r")
    if r is not None:
        t = r.makeelement(W + "t", {})
        t.set(XMLSPACE, "preserve")
        t.text = valeur
        r.append(t)
        return True
    return False


def traiter_partie(element, vals, compteur):
    for sdt in element.iter(W + "sdt"):
        pr = sdt.find(W + "sdtPr")
        if pr is None:
            continue
        tag_el = pr.find(W + "tag")
        if tag_el is None:
            continue
        tag = tag_el.get(W + "val")
        if tag in vals and remplir_sdt(sdt, vals[tag]):
            compteur[tag] = compteur.get(tag, 0) + 1


def main():
    root = tk.Tk()
    root.withdraw()

    docx_in = filedialog.askopenfilename(
        title="1/2 — Choisissez le MODELE Word (.docx)",
        filetypes=[("Document Word", "*.docx")])
    if not docx_in:
        return
    xlsx = filedialog.askopenfilename(
        title="2/2 — Choisissez le TABLEUR de valeurs (.xlsx)",
        filetypes=[("Classeur Excel", "*.xlsx")])
    if not xlsx:
        return

    try:
        vals = lire_valeurs(xlsx)
        if not vals:
            messagebox.showerror("Erreur",
                "Aucune valeur trouvée. Vérifiez que le tableur a bien "
                "une colonne 'Tag' et une colonne 'Valeur' remplie.")
            return

        doc = Document(docx_in)
        compteur = {}
        traiter_partie(doc.element.body, vals, compteur)
        for sec in doc.sections:
            for part in (sec.header, sec.footer,
                         sec.first_page_header, sec.first_page_footer,
                         sec.even_page_header, sec.even_page_footer):
                if part is not None:
                    traiter_partie(part._element, vals, compteur)

        base, ext = os.path.splitext(docx_in)
        sortie = base + "_rempli.docx"
        doc.save(sortie)

        total = sum(compteur.values())
        manquants = [t for t in vals if t not in compteur]
        msg = f"Terminé !\n\n{total} contrôle(s) rempli(s) pour {len(compteur)} tag(s).\n\nFichier créé :\n{sortie}"
        if manquants:
            msg += "\n\nTags renseignés mais absents du document :\n - " + "\n - ".join(manquants)
        messagebox.showinfo("Remplissage FDES", msg)

    except Exception as e:
        messagebox.showerror("Erreur", f"Un problème est survenu :\n\n{e}")


if __name__ == "__main__":
    main()